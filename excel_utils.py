import openpyxl
from helpers import format_headers, format_values

def read_excel(file_path: str, extra_columns_flag: bool, num_extra_cols: int) -> dict[str, dict[str, list[str]]]:
    """
    Reads an Excel file and extracts headers from the first column and values from the second column
    for all sheets. Returns a dictionary with sheet names as keys and extracted data as values.
    :param file_path: Absolute path to file
    :param extra_columns_flag: Flag to read and parse extra lines of data
    :param num_extra_cols: Defines how many extra lines of data to parse
    """
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        headers = [row[0] for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True
        )]

        values = [[row[0] for row in sheet.iter_rows(
                    min_row=1, max_row=sheet.max_row, min_col=2, max_col=2, values_only=True
                )]]

        if extra_columns_flag:
            for col in range(3, 3+num_extra_cols):
                values.append([
                    row[0] for row in sheet.iter_rows(
                        min_row=1, max_row=sheet.max_row, min_col=col, max_col=col, values_only=True
                    )
                ])

        try:
            start_index = values[0].index(1) + 1
            filtered_headers = format_headers(headers[start_index:])
            # if extra_columns_flag:
            #     filtered_headers.insert(0, "Subsets")
            #     filtered_values = [value[start_index - 2:] for value in values]
            #     filtered_values = [value[:1] + value[2:] for value in filtered_values]
            # else:
            #     filtered_values = [value[start_index:] for value in values]
            filtered_values = [value[start_index:] for value in values]
            filtered_values = [format_values(value) for value in filtered_values]

        except ValueError:
            filtered_headers = []
            filtered_values = []

        data[sheet_name] = {
            "headers": filtered_headers,
            "values": filtered_values
        }

    return data


def get_question_data(file_path: str) -> list[list[str]]:
    """
    Reads an Excel file and extracts the prerequisite data on each sheet
    :param file_path: Absolute path to file
    """
    workbook = openpyxl.load_workbook(file_path)
    pre_data = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        headers = [row[0] for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True
        )]

        try:
            # Magic line - Essentially searches through the list by enumerating the list, then checking the instance
            # to make sure it's a string, then searching to see if it contains the string
            end_index = next(
                (i for i, header in enumerate(headers) if isinstance(header, str) and "BASE=" in header), -1) - 1
            filtered_headers = headers[:end_index]
        except ValueError:
            filtered_headers = []

        pre_data.append(filtered_headers)

    return pre_data



